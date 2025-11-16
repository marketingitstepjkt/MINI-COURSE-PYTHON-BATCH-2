import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# ---------------------------
# Helper Functions
# ---------------------------
def wait_and_find(driver, locator, timeout=10):
    """Tunggu elemen sampai terlihat lalu return element."""
    return WebDriverWait(driver, timeout).until(EC.visibility_of_element_located(locator))


def fill_input(driver, locator, value):
    """Isi input field dengan aman."""
    element = wait_and_find(driver, locator)
    element.clear()
    element.send_keys(str(value))


def is_row_valid(row):
    """Cek apakah semua kolom pada 1 baris TERISI (tidak boleh None atau kosong)."""
    return all(cell not in (None, "", " ") for cell in row)


# ---------------------------
# Main Program
# ---------------------------

def main():
    # Load Excel
    wb = load_workbook("Day 2/data.xlsx")
    sheet = wb["Sheet1"]

    # Setup Browser
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://demoqa.com/webtables")
    driver.implicitly_wait(10)

    print("Automation starting...\n")

    # Loop semua baris data mulai dari baris ke-2 (skip header)
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

        # Validasi data harus lengkap
        if not is_row_valid(row):
            print(f"❌ Baris {row_index} dilewati karena ada value yang kosong: {row}")
            continue

        # Ambil data
        first_name, last_name, age, email, salary, department = row

        # Klik tombol "Add"
        wait_and_find(driver, (By.ID, "addNewRecordButton")).click()

        try:
            # Tunggu form muncul
            wait_and_find(driver, (By.ID, "registration-form-modal"))

            # Isi input
            fill_input(driver, (By.ID, "firstName"), first_name)
            fill_input(driver, (By.ID, "lastName"), last_name)
            fill_input(driver, (By.ID, "userEmail"), email)
            fill_input(driver, (By.ID, "age"), age)
            fill_input(driver, (By.ID, "salary"), salary)
            fill_input(driver, (By.ID, "department"), department)

            # Submit form
            wait_and_find(driver, (By.ID, "submit")).click()

        except (TimeoutException, NoSuchElementException):
            print(f"⚠ Error pada baris {row_index}. Data gagal diinput.")
            continue

        print(f"✔ Baris {row_index} berhasil diinput: {first_name} {last_name}")
        time.sleep(1)

    print("\n🎉 Semua data valid berhasil diinput!")
    driver.quit()


# Start program
if __name__ == "__main__":
    main()
